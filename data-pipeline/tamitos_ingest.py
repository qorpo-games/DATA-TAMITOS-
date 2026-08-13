#!/usr/bin/env python3
"""
TAMITOS — proof-of-concept zberu dát do jednej databázy.

Stiahne dva verejné zdroje a zjednotí ich do jednej SQLite databázy:
  1) ClinicalTrials.gov v2 API  -> klinické štúdie o autizme (tabuľka `studies`)
  2) CVTI adresár CPP/ŠCPP (XLSX) -> poradenské centrá na SK (tabuľka `providers`)

Princíp, ktorý chráni projekt: ťaháme len z PRIMÁRNYCH verejných zdrojov
(štátne registre, oficiálne API) — nie z cudzích zozbieraných databáz.

Spustenie:
    python3 tamitos_ingest.py                 # naostro (potrebný internet)
    python3 tamitos_ingest.py --sample         # offline ukážka (bundled dáta)
    python3 tamitos_ingest.py --db tamitos.db  # vlastný súbor DB

Určené na denné spúšťanie cez cron (napr. 05:00) — je idempotentné
(UPSERT podľa prirodzeného kľúča), takže opakované behy dáta neduplikujú.
"""
import argparse, json, sqlite3, sys, datetime, urllib.request, urllib.parse, io, os

CT_API = "https://clinicaltrials.gov/api/v2/studies"
# CVTI adresár CPP/ŠCPP (verejný XLSX). URL sa občas mení — over na cvtisr.sk.
CVTI_XLSX = "https://www.cvtisr.sk/buxus/docs//prevencia/CPP_SCPP/Adresar_organizacii_CPP_SCPP_2.xlsx"

SAMPLE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sample_data")


# ---------------------------------------------------------------- DB schema
def init_db(path):
    con = sqlite3.connect(path)
    con.executescript("""
    CREATE TABLE IF NOT EXISTS studies (
        nct_id       TEXT PRIMARY KEY,          -- prirodzený kľúč
        title        TEXT,
        status       TEXT,
        phase        TEXT,
        conditions   TEXT,
        interventions TEXT,
        countries    TEXT,
        has_slovakia INTEGER,
        url          TEXT,
        source       TEXT,
        fetched_at   TEXT
    );
    CREATE TABLE IF NOT EXISTS providers (
        ext_id       TEXT PRIMARY KEY,          -- zdroj:riadok / IČO / názov+mesto
        name         TEXT,
        kind         TEXT,                      -- typ zariadenia
        region       TEXT,
        city         TEXT,
        address      TEXT,
        contact      TEXT,
        source       TEXT,
        fetched_at   TEXT
    );
    CREATE TABLE IF NOT EXISTS ingest_log (
        run_at TEXT, source TEXT, rows INTEGER, note TEXT
    );
    """)
    con.commit()
    return con


def log(con, source, rows, note=""):
    con.execute("INSERT INTO ingest_log VALUES (?,?,?,?)",
                (datetime.datetime.now(datetime.timezone.utc).isoformat(timespec="seconds"), source, rows, note))
    con.commit()


# ---------------------------------------------------- 1) ClinicalTrials.gov
def fetch_clinicaltrials(cond="autism", locn=None, sample=False):
    if sample:
        with open(os.path.join(SAMPLE_DIR, "clinicaltrials_sample.json"), encoding="utf-8") as f:
            return json.load(f).get("studies", [])
    params = {"query.cond": cond, "pageSize": 100}
    if locn:
        params["query.locn"] = locn
    out, token = [], None
    while True:
        if token:
            params["pageToken"] = token
        url = CT_API + "?" + urllib.parse.urlencode(params)
        req = urllib.request.Request(url, headers={"User-Agent": "TAMITOS-ingest/1.0"})
        with urllib.request.urlopen(req, timeout=30) as r:
            data = json.load(r)
        out.extend(data.get("studies", []))
        token = data.get("nextPageToken")
        if not token:
            break
    return out


def normalize_study(s):
    ps = s.get("protocolSection", {})
    idm = ps.get("identificationModule", {})
    status = ps.get("statusModule", {}).get("overallStatus")
    design = ps.get("designModule", {})
    phases = ", ".join(design.get("phases", []) or [])
    conds = ", ".join(ps.get("conditionsModule", {}).get("conditions", []) or [])
    ints = ", ".join(i.get("name", "") for i in ps.get("armsInterventionsModule", {}).get("interventions", []) or [])
    locs = ps.get("contactsLocationsModule", {}).get("locations", []) or []
    countries = sorted({l.get("country") for l in locs if l.get("country")})
    nct = idm.get("nctId")
    return {
        "nct_id": nct,
        "title": idm.get("briefTitle"),
        "status": status,
        "phase": phases,
        "conditions": conds,
        "interventions": ints,
        "countries": ", ".join(countries),
        "has_slovakia": int("Slovakia" in countries),
        "url": f"https://clinicaltrials.gov/study/{nct}" if nct else None,
        "source": "ClinicalTrials.gov v2 API",
        "fetched_at": datetime.datetime.now(datetime.timezone.utc).isoformat(timespec="seconds"),
    }


def upsert_studies(con, rows):
    con.executemany("""
        INSERT INTO studies VALUES
        (:nct_id,:title,:status,:phase,:conditions,:interventions,:countries,:has_slovakia,:url,:source,:fetched_at)
        ON CONFLICT(nct_id) DO UPDATE SET
          title=excluded.title, status=excluded.status, phase=excluded.phase,
          conditions=excluded.conditions, interventions=excluded.interventions,
          countries=excluded.countries, has_slovakia=excluded.has_slovakia,
          url=excluded.url, fetched_at=excluded.fetched_at
    """, rows)
    con.commit()


# ------------------------------------------------------- 2) CVTI XLSX adresár
def load_cvti_xlsx(sample=False):
    import openpyxl
    if sample:
        path = os.path.join(SAMPLE_DIR, "cvti_sample.xlsx")
        if not os.path.exists(path):  # binárka nie je v gite – vygeneruj ju
            import subprocess
            subprocess.run([sys.executable, os.path.join(SAMPLE_DIR, "make_sample_xlsx.py")], check=True)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    else:
        req = urllib.request.Request(CVTI_XLSX, headers={"User-Agent": "TAMITOS-ingest/1.0"})
        with urllib.request.urlopen(req, timeout=60) as r:
            wb = openpyxl.load_workbook(io.BytesIO(r.read()), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip().lower() if h else "" for h in rows[0]]

    def col(*names):
        for n in names:
            for i, h in enumerate(header):
                if n in h:
                    return i
        return None

    ci = {
        "name": col("názov", "nazov", "organiz"),
        "kind": col("typ", "druh", "zariaden"),
        "region": col("kraj"),
        "city": col("mesto", "obec"),
        "address": col("adresa", "ulica"),
        "contact": col("mail", "e-mail", "telï¿½fón", "telefon", "kontakt"),
    }
    out = []
    for idx, r in enumerate(rows[1:], start=2):
        def get(k):
            return str(r[ci[k]]).strip() if ci[k] is not None and r[ci[k]] is not None else ""
        name = get("name")
        if not name:
            continue
        out.append({
            "ext_id": f"cvti:{idx}",
            "name": name,
            "kind": get("kind") or "CPP/ŠCPP",
            "region": get("region"),
            "city": get("city"),
            "address": get("address"),
            "contact": get("contact"),
            "source": "CVTI adresár CPP/ŠCPP",
            "fetched_at": datetime.datetime.now(datetime.timezone.utc).isoformat(timespec="seconds"),
        })
    return out


def upsert_providers(con, rows):
    con.executemany("""
        INSERT INTO providers VALUES
        (:ext_id,:name,:kind,:region,:city,:address,:contact,:source,:fetched_at)
        ON CONFLICT(ext_id) DO UPDATE SET
          name=excluded.name, kind=excluded.kind, region=excluded.region,
          city=excluded.city, address=excluded.address, contact=excluded.contact,
          fetched_at=excluded.fetched_at
    """, rows)
    con.commit()


# --------------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser(description="TAMITOS data ingest PoC")
    ap.add_argument("--db", default="tamitos.db")
    ap.add_argument("--sample", action="store_true", help="offline režim – bundled ukážkové dáta")
    ap.add_argument("--locn", default=None, help="filter lokality štúdií, napr. Slovakia")
    args = ap.parse_args()

    con = init_db(args.db)
    print(f"→ databáza: {args.db}  ({'SAMPLE' if args.sample else 'LIVE'} režim)\n")

    # 1) štúdie
    try:
        raw = fetch_clinicaltrials(locn=args.locn, sample=args.sample)
        studies = [normalize_study(s) for s in raw if s.get("protocolSection")]
        upsert_studies(con, studies)
        sk = sum(s["has_slovakia"] for s in studies)
        log(con, "ClinicalTrials.gov", len(studies), f"{sk} so SK lokalitou")
        print(f"✓ ClinicalTrials.gov: {len(studies)} štúdií (z toho {sk} so slovenskou lokalitou)")
    except Exception as e:
        print(f"✗ ClinicalTrials.gov zlyhalo: {e}", file=sys.stderr)

    # 2) poskytovatelia
    try:
        provs = load_cvti_xlsx(sample=args.sample)
        upsert_providers(con, provs)
        log(con, "CVTI CPP/ŠCPP", len(provs))
        print(f"✓ CVTI adresár CPP/ŠCPP: {len(provs)} poskytovateľov")
    except Exception as e:
        print(f"✗ CVTI XLSX zlyhalo: {e}", file=sys.stderr)

    # súhrn
    st = con.execute("SELECT COUNT(*) FROM studies").fetchone()[0]
    pr = con.execute("SELECT COUNT(*) FROM providers").fetchone()[0]
    print(f"\n── V databáze spolu: {st} štúdií, {pr} poskytovateľov ──")
    print("\nUkážka – štúdie so slovenskou lokalitou:")
    for row in con.execute("SELECT nct_id,status,title FROM studies WHERE has_slovakia=1 LIMIT 5"):
        print(f"   [{row[1]:<18}] {row[0]}  {row[2][:60]}")
    print("\nUkážka – poskytovatelia podľa kraja:")
    for row in con.execute("SELECT region,COUNT(*) FROM providers GROUP BY region ORDER BY 2 DESC LIMIT 8"):
        print(f"   {row[0] or '(neuvedené)':<18} {row[1]}")
    con.close()


if __name__ == "__main__":
    main()
